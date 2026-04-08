from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _autosize(ws, max_width: int = 60):
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            s = str(v)
            widths[i] = max(widths.get(i, 0), min(max_width, len(s) + 2))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def write_workbook(
    diff_rows: list[dict[str, Any]],
    peer_rows: list[dict[str, Any]],
    risk_rows: list[dict[str, Any]],
    out_path: str | Path,
):
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    def write_sheet(name: str, rows: list[dict[str, Any]]):
        ws = wb.create_sheet(title=name)
        if not rows:
            ws.append(["(no rows)"])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h) for h in headers])
        _autosize(ws)

    write_sheet("DIFF_TABLE", diff_rows)
    write_sheet("PEER_BENCHMARK", peer_rows)
    write_sheet("RISK_COMMENTS", risk_rows)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

